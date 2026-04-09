#!/usr/bin/env python3
"""
Extrahiert die Gold-Standard-Referenzwerte aus Riegelsberg LV3.xlsx.

Wird manuell aufgerufen, wenn die Quell-Datei aktualisiert wurde.
Schreibt `test-daten/riegelsberg/referenz.json`.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl fehlt. Installiere mit: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

QUELL_DATEI = Path(
    "/Users/admin/Library/CloudStorage/OneDrive-kalku/KT01/"
    "1695_Gesellchen_GmbH/_abgeschlossen/260319_Friedhoefe_Riegelsberg/LV3.xlsx"
)

ZIEL_DATEI = Path(__file__).resolve().parent.parent / "test-daten" / "riegelsberg" / "referenz.json"


def main() -> int:
    if not QUELL_DATEI.exists():
        print(f"ERROR: Quell-Datei nicht gefunden: {QUELL_DATEI}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(str(QUELL_DATEI), data_only=True)
    ws = wb["Kalkulation"]

    parameter = {
        "verrechnungslohn": ws["M2"].value,
        "lohn_ek": ws["K2"].value,
        "materialzuschlag": ws["K4"].value,
        "nzuschlag": ws["K5"].value,
        "geraete_grundzuschlag": ws["K6"].value,
        "zeitwert_faktor": ws["AP5"].value,
        "geraetezulage_default": ws["AP3"].value,
    }

    positionen = []
    for r in range(15, ws.max_row + 1):
        oz = ws.cell(row=r, column=1).value
        if oz is None:
            continue
        menge = ws.cell(row=r, column=3).value
        einheit = ws.cell(row=r, column=4).value
        kurztext = ws.cell(row=r, column=2).value
        if kurztext:
            kurztext = str(kurztext).split("\n")[0].strip()
        ep = ws.cell(row=r, column=5).value
        gp = ws.cell(row=r, column=6).value
        if ep is None or menge is None or einheit is None:
            continue

        positionen.append({
            "oz": str(oz).strip(),
            "kurztext": kurztext,
            "menge": menge,
            "einheit": str(einheit),
            "X_stoffe_ek": ws.cell(row=r, column=24).value,
            "Y_zeit_min_roh": ws.cell(row=r, column=25).value,
            "Z_geraete_eur_h": ws.cell(row=r, column=26).value,
            "M_nu_ek": ws.cell(row=r, column=13).value,
            "AC_zeit_mit_faktor": ws.cell(row=r, column=29).value,
            "AA_geraete_ep": ws.cell(row=r, column=27).value,
            "AB_lohn_ep": ws.cell(row=r, column=28).value,
            "AJ_stoffe_vk": ws.cell(row=r, column=36).value,
            "AK_nu_vk": ws.cell(row=r, column=37).value,
            "EP": ep,
            "GP": gp,
        })

    ZIEL_DATEI.parent.mkdir(parents=True, exist_ok=True)
    ZIEL_DATEI.write_text(
        json.dumps(
            {"parameter": parameter, "positionen": positionen},
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    print(f"OK: {len(positionen)} Positionen → {ZIEL_DATEI}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
